"""
ViewSets for Asset and related models.

Provides:
- AssetViewSet: Asset CRUD with status change and statistics
- SupplierViewSet: Supplier CRUD
- LocationViewSet: Location CRUD with tree support
- AssetStatusLogViewSet: Status log viewing
"""
import re
from django.conf import settings
from rest_framework.decorators import action
from rest_framework.response import Response
from rest_framework import status
from django.http import HttpResponse
from apps.common.viewsets.base import BaseModelViewSetWithBatch
from apps.common.responses.base import BaseResponse
from apps.assets.models import Asset, Supplier, Location, AssetStatusLog
from apps.assets.serializers import (
    AssetListSerializer,
    AssetDetailSerializer,
    AssetCreateSerializer,
    AssetUpdateSerializer,
    AssetStatusSerializer,
    AssetSerializer,
    SupplierSerializer,
    SupplierListSerializer,
    LocationSerializer,
    LocationListSerializer,
    AssetStatusLogSerializer,
    AssetStatusLogListSerializer,
)
from apps.assets.filters import AssetFilter, SupplierFilter, LocationFilter, AssetStatusLogFilter
from apps.assets.services import AssetService, SupplierService, LocationService, AssetStatusLogService

# Maximum number of QR codes that can be generated in a single bulk request
MAX_BULK_QR_LIMIT = 1000


def sanitize_filename(name: str) -> str:
    r"""
    Sanitize a string to be safe for use as a filename.

    Removes or replaces characters that are invalid in Windows filenames:
    < > : " / \ | ? *

    Args:
        name: The string to sanitize

    Returns:
        A sanitized string safe for use as a filename
    """
    # Replace invalid filename characters with underscore
    invalid_chars = r'[<>:"/\\|?*]'
    return re.sub(invalid_chars, '_', name)


class AssetViewSet(BaseModelViewSetWithBatch):
    """
    ViewSet for Asset management.

    Provides:
    - Standard CRUD operations
    - Status change with audit logging
    - QR/RFID code lookup
    - Asset statistics
    - Bulk import/export
    """

    queryset = Asset.objects.select_related(
        'asset_category',
        'supplier',
        'department',
        'location',
        'custodian',
        'user',
        'created_by'
    ).all()
    serializer_class = AssetSerializer
    filterset_class = AssetFilter
    service = AssetService()

    # Search fields for search query parameter
    search_fields = ['asset_code', 'asset_name', 'specification', 'brand', 'model', 'serial_number']

    def get_serializer_class(self):
        """Return appropriate serializer based on action."""
        if self.action == 'list':
            return AssetListSerializer
        if self.action == 'retrieve':
            return AssetDetailSerializer
        if self.action == 'create':
            return AssetCreateSerializer
        if self.action == 'update':
            return AssetUpdateSerializer
        return super().get_serializer_class()

    def get_serializer_context(self):
        """Add organization_id to serializer context."""
        context = super().get_serializer_context()
        # Try to get organization_id from request attribute (set by middleware)
        if hasattr(self.request, 'organization_id') and self.request.organization_id:
            context['organization_id'] = str(self.request.organization_id)
        # Fallback to user's organization_id if available
        elif hasattr(self.request, 'user') and hasattr(self.request.user, 'organization_id'):
            context['organization_id'] = str(self.request.user.organization_id)
        return context

    def perform_create(self, serializer):
        """Set organization and created_by on create."""
        organization_id = getattr(self.request, 'organization_id', None)
        serializer.save(
            organization_id=organization_id,
            created_by=self.request.user
        )

    def create(self, request, *args, **kwargs):
        """Override to wrap response in standard format."""
        serializer = self.get_serializer(data=request.data)
        serializer.is_valid(raise_exception=True)
        self.perform_create(serializer)
        headers = self.get_success_headers(serializer.data)

        return BaseResponse.created(
            data=serializer.data,
            message='Asset created successfully'
        )

    def retrieve(self, request, *args, **kwargs):
        """Override to wrap response in standard format."""
        instance = self.get_object()
        serializer = self.get_serializer(instance)
        return BaseResponse.success(
                data=serializer.data,
                message='Query successful'
            )

    def update(self, request, *args, **kwargs):
        """Override to wrap response in standard format."""
        partial = kwargs.pop('partial', False)
        instance = self.get_object()
        serializer = self.get_serializer(
            instance, data=request.data, partial=partial
        )
        serializer.is_valid(raise_exception=True)
        self.perform_update(serializer)

        return BaseResponse.success(
                data=serializer.data,
                message='Asset updated successfully'
            )

    def destroy(self, request, *args, **kwargs):
        """Override to wrap response in standard format."""
        instance = self.get_object()
        self.perform_destroy(instance)
        return BaseResponse.success(message='Asset deleted successfully')

    @action(detail=False, methods=['get'], url_path='statistics')
    def statistics(self, request):
        """
        Get asset statistics for current organization.

        GET /api/assets/statistics/

        Returns:
        {
            "success": true,
            "data": {
                "total": 100,
                "by_status": {...},
                "total_value": 1000000.00,
                "total_net_value": 800000.00,
                "by_category": {...}
            }
        }
        """
        organization_id = getattr(request, 'organization_id', None)

        if not organization_id:
            return BaseResponse.error(
                    code='UNAUTHORIZED',
                    message='Organization context required'
                )

        stats = self.service.get_asset_statistics(organization_id)

        return BaseResponse.success(
                data=stats,
                message='Statistics retrieved successfully'
            )

    @action(detail=False, methods=['get'], url_path='depreciation-summary')
    def depreciation_summary(self, request):
        """
        Get depreciation summary for current organization.

        GET /api/assets/depreciation-summary/
        """
        organization_id = getattr(request, 'organization_id', None)

        if not organization_id:
            return BaseResponse.error(
                    code='UNAUTHORIZED',
                    message='Organization context required'
                )

        summary = self.service.get_depreciation_summary(organization_id)

        return BaseResponse.success(
                data=summary,
                message='Depreciation summary retrieved successfully'
            )

    @action(detail=True, methods=['post'], url_path='change-status')
    def change_status(self, request, pk=None):
        """
        Change asset status with audit logging.

        POST /api/assets/{id}/change_status/
        Body:
        {
            "new_status": "in_use",
            "reason": "Asset assigned to user"
        }
        """
        asset = self.get_object()
        serializer = AssetStatusSerializer(data=request.data)

        serializer.is_valid(raise_exception=True)

        try:
            updated_asset = self.service.change_status(
                asset_id=asset.id,
                new_status=serializer.validated_data['new_status'],
                user=request.user,
                reason=serializer.validated_data.get('reason', '')
            )
        except ValueError as e:
            return BaseResponse.error(
                code='VALIDATION_ERROR',
                message=str(e)
            )

        response_serializer = AssetDetailSerializer(updated_asset)

        return BaseResponse.success(
                data=response_serializer.data,
                message='Asset status changed successfully'
            )

    @action(detail=False, methods=['get'], url_path='lookup')
    def lookup(self, request):
        """
        Lookup asset by QR code or RFID code.

        GET /api/assets/lookup/?qr_code=xxx or ?rfid_code=xxx
        """
        qr_code = request.query_params.get('qr_code')
        rfid_code = request.query_params.get('rfid_code')

        if not qr_code and not rfid_code:
            return BaseResponse.error(
                    code='VALIDATION_ERROR',
                    message='Either qr_code or rfid_code parameter is required'
                )

        organization_id = getattr(request, 'organization_id', None)
        asset = None

        if qr_code:
            asset = self.service.get_by_qr_code(qr_code, organization_id)
        elif rfid_code:
            asset = self.service.get_by_rfid_code(rfid_code, organization_id)

        if not asset:
            return BaseResponse.error(
                    code='NOT_FOUND',
                    message='Asset not found'
                )

        serializer = AssetDetailSerializer(asset)
        return BaseResponse.success(
                data=serializer.data,
                message='Asset found'
            )

    @action(detail=False, methods=['get'], url_path='by-category/(?P<category_id>[^/.]+)')
    def by_category(self, request, category_id=None):
        """
        Get assets by category.

        GET /api/assets/by-category/{category_id}/?include_children=true
        """
        organization_id = getattr(request, 'organization_id', None)
        include_children = request.query_params.get('include_children', 'true').lower() == 'true'

        queryset = self.service.query_by_category(
            category_id=category_id,
            organization_id=organization_id,
            include_children=include_children
        )

        page = self.paginate_queryset(queryset)
        if page is not None:
            serializer = AssetListSerializer(page, many=True)
            return self.get_paginated_response(serializer.data)

        serializer = AssetListSerializer(queryset, many=True)
        return BaseResponse.success(
                data=serializer.data,
                message='Query successful'
            )

    @action(detail=False, methods=['get'], url_path='by-department/(?P<department_id>[^/.]+)')
    def by_department(self, request, department_id=None):
        """
        Get assets by department.

        GET /api/assets/by-department/{department_id}/
        """
        organization_id = getattr(request, 'organization_id', None)

        queryset = self.service.query_by_department(
            department_id=department_id,
            organization_id=organization_id
        )

        page = self.paginate_queryset(queryset)
        if page is not None:
            serializer = AssetListSerializer(page, many=True)
            return self.get_paginated_response(serializer.data)

        serializer = AssetListSerializer(queryset, many=True)
        return BaseResponse.success(
                data=serializer.data,
                message='Query successful'
            )

    @action(detail=False, methods=['get'], url_path='by-location/(?P<location_id>[^/.]+)')
    def by_location(self, request, location_id=None):
        """
        Get assets by location.

        GET /api/assets/by-location/{location_id}/?include_children=true
        """
        organization_id = getattr(request, 'organization_id', None)
        include_children = request.query_params.get('include_children', 'true').lower() == 'true'

        queryset = self.service.query_by_location(
            location_id=location_id,
            organization_id=organization_id,
            include_children=include_children
        )

        page = self.paginate_queryset(queryset)
        if page is not None:
            serializer = AssetListSerializer(page, many=True)
            return self.get_paginated_response(serializer.data)

        serializer = AssetListSerializer(queryset, many=True)
        return BaseResponse.success(
                data=serializer.data,
                message='Query successful'
            )

    @action(detail=True, methods=['get'], url_path='status-history')
    def status_history(self, request, pk=None):
        """
        Get status change history for an asset.

        GET /api/assets/{id}/status-history/
        """
        asset = self.get_object()
        logs = self.service.status_log_service.get_asset_history(asset.id)

        page = self.paginate_queryset(logs)
        if page is not None:
            serializer = AssetStatusLogListSerializer(page, many=True)
            return self.get_paginated_response(serializer.data)

        serializer = AssetStatusLogListSerializer(logs, many=True)
        return BaseResponse.success(
                data=serializer.data,
                message='Status history retrieved successfully'
            )

    @action(detail=True, methods=['get'], url_path='qr_code')
    def qr_code(self, request, pk=None):
        """
        Generate QR code for an asset.

        GET /api/assets/{id}/qr_code/

        Returns a PNG image containing the QR code.
        The QR code contains the asset URL for quick scanning.
        """
        import qrcode
        from io import BytesIO

        asset = self.get_object()

        # Build QR code data - asset detail URL
        frontend_url = getattr(settings, 'FRONTEND_URL', 'http://localhost:5173')
        qr_data = f'{frontend_url}/assets/{asset.id}'

        # Generate QR code
        qr = qrcode.QRCode(
            version=1,
            error_correction=qrcode.constants.ERROR_CORRECT_L,
            box_size=10,
            border=4,
        )
        qr.add_data(qr_data)
        qr.make(fit=True)

        # Create image
        img = qr.make_image(fill_color='black', back_color='white')

        # Convert to bytes
        buffer = BytesIO()
        img.save(buffer, format='PNG')
        buffer.seek(0)

        # Return as image response
        return HttpResponse(
            buffer.read(),
            content_type='image/png'
        )

    @action(detail=False, methods=['post'], url_path='bulk-qr-codes')
    def bulk_qr_codes(self, request):
        """
        Generate QR codes for multiple assets as a ZIP file.

        POST /api/assets/bulk-qr-codes/
        Body:
        {
            "ids": ["uuid1", "uuid2", "uuid3"]
        }

        Returns a ZIP file containing PNG images.

        Validation:
        - ids must be a list
        - ids cannot exceed MAX_BULK_QR_LIMIT (1000)
        - ids cannot be empty
        """
        import qrcode
        from io import BytesIO
        from zipfile import ZipFile

        ids = request.data.get('ids', [])

        # Validate that ids is a list
        if not isinstance(ids, list):
            return BaseResponse.error(
                code='VALIDATION_ERROR',
                message='ids parameter must be a list'
            )

        # Validate that ids is not empty
        if not ids:
            return BaseResponse.error(
                code='VALIDATION_ERROR',
                message='ids parameter cannot be empty'
            )

        # Validate maximum limit
        if len(ids) > MAX_BULK_QR_LIMIT:
            return BaseResponse.error(
                code='VALIDATION_ERROR',
                message=f'Cannot generate more than {MAX_BULK_QR_LIMIT} QR codes at once'
            )

        organization_id = getattr(request, 'organization_id', None)
        assets = Asset.objects.filter(
            id__in=ids,
            organization_id=organization_id,
            is_deleted=False
        )

        # Create ZIP file in memory
        zip_buffer = BytesIO()
        frontend_url = getattr(settings, 'FRONTEND_URL', 'http://localhost:5173')

        # Create reusable QRCode instance for better performance
        qr_maker = qrcode.QRCode(
            version=1,
            error_correction=qrcode.constants.ERROR_CORRECT_L,
            box_size=10,
            border=4,
        )

        with ZipFile(zip_buffer, 'w') as zip_file:
            for asset in assets:
                # Sanitize asset code for safe filename usage
                safe_code = sanitize_filename(asset.asset_code or str(asset.id))
                qr_data = f'{frontend_url}/assets/{asset.id}'

                # Clear and reuse QRCode instance
                qr_maker.clear()
                qr_maker.add_data(qr_data)
                qr_maker.make(fit=True)
                img = qr_maker.make_image(fill_color='black', back_color='white')

                img_buffer = BytesIO()
                img.save(img_buffer, format='PNG')
                img_buffer.seek(0)

                filename = f"QR_{safe_code}.png"
                zip_file.writestr(filename, img_buffer.read())

        zip_buffer.seek(0)

        return HttpResponse(
            zip_buffer.read(),
            content_type='application/zip',
            headers={
                'Content-Disposition': 'attachment; filename="asset_qr_codes.zip"'
            }
        )

    @action(detail=False, methods=['post'], url_path='bulk-import')
    def bulk_import(self, request):
        """
        Bulk import assets.

        POST /api/assets/bulk-import/
        Body:
        {
            "assets": [
                {"asset_name": "...", "asset_category_id": "...", ...},
                ...
            ]
        }
        """
        from apps.assets.serializers import AssetBulkImportSerializer

        serializer = AssetBulkImportSerializer(data=request.data)
        serializer.is_valid(raise_exception=True)

        organization_id = getattr(request, 'organization_id', None)

        result = self.service.bulk_create(
            assets_data=serializer.validated_data['assets'],
            user=request.user,
            organization_id=organization_id
        )

        http_status = status.HTTP_200_OK if result['failed'] == 0 else status.HTTP_207_MULTI_STATUS

        return BaseResponse.success(
            data=result,
            message=f'Bulk import completed: {result["succeeded"]} succeeded, {result["failed"]} failed',
            http_status=http_status
        )

    @action(detail=False, methods=['post'], url_path='batch_change_status')
    def batch_change_status(self, request):
        """
        Batch change asset status.

        POST /api/assets/batch_change_status/
        Body:
        {
            "ids": ["uuid1", "uuid2", ...],
            "new_status": "maintenance"
        }
        """
        ids = request.data.get('ids', [])
        new_status = request.data.get('new_status')

        if not ids:
            return Response(
                {'success': False, 'error': {'code': 'MISSING_IDS', 'message': 'IDs are required'}},
                status=status.HTTP_400_BAD_REQUEST
            )

        if not new_status:
            return Response(
                {'success': False, 'error': {'code': 'MISSING_STATUS', 'message': 'New status is required'}},
                status=status.HTTP_400_BAD_REQUEST
            )

        results = []
        succeeded = 0
        failed = 0
        organization_id = getattr(request, 'organization_id', None)

        for asset_id in ids:
            try:
                asset = Asset.objects.get(id=asset_id, organization_id=organization_id)
                asset.asset_status = new_status
                asset.save()
                succeeded += 1
                results.append({'id': str(asset_id), 'success': True})
            except Asset.DoesNotExist:
                failed += 1
                results.append({'id': str(asset_id), 'success': False, 'error': 'Asset not found'})

        return Response({
            'success': True,
            'message': f'Batch status change completed',
            'summary': {
                'total': len(ids),
                'succeeded': succeeded,
                'failed': failed
            },
            'results': results
        })

    @action(detail=False, methods=['post'], url_path='export')
    def export_assets(self, request):
        """
        Export assets to Excel file.

        POST /api/assets/export/
        Body:
        {
            "filters": {...},  // Optional filters
            "columns": ["asset_code", "asset_name", ...]  // Columns to export
        }

        Returns an Excel file (.xlsx) with the filtered assets.
        """
        import openpyxl
        from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
        from datetime import datetime
        from io import BytesIO
        from django.core.paginator import Paginator

        # Get request parameters
        filters = request.data.get('filters', {})
        columns = request.data.get('columns', [])

        # Default columns if not specified
        if not columns:
            columns = [
                'asset_code', 'asset_name', 'specification', 'brand', 'model',
                'serial_number', 'asset_category', 'supplier', 'department',
                'location', 'custodian', 'asset_status', 'purchase_date',
                'original_value', 'net_value', 'useful_life', 'created_at'
            ]

        # Column headers mapping
        column_headers = {
            'asset_code': 'Asset Code',
            'asset_name': 'Asset Name',
            'specification': 'Specification',
            'brand': 'Brand',
            'model': 'Model',
            'serial_number': 'Serial Number',
            'asset_category': 'Category',
            'supplier': 'Supplier',
            'department': 'Department',
            'location': 'Location',
            'custodian': 'Custodian',
            'asset_status': 'Status',
            'purchase_date': 'Purchase Date',
            'original_value': 'Original Value',
            'net_value': 'Net Value',
            'useful_life': 'Useful Life (months)',
            'created_at': 'Created At'
        }

        # Get base queryset
        organization_id = getattr(request, 'organization_id', None)
        queryset = Asset.objects.filter(
            organization_id=organization_id,
            is_deleted=False
        ).select_related(
            'asset_category', 'supplier', 'department', 'location', 'custodian'
        )

        # Apply filters if provided
        if filters:
            # Filter by status
            if 'asset_status' in filters:
                queryset = queryset.filter(asset_status=filters['asset_status'])

            # Filter by category
            if 'asset_category_id' in filters:
                queryset = queryset.filter(asset_category_id=filters['asset_category_id'])

            # Filter by department
            if 'department_id' in filters:
                queryset = queryset.filter(department_id=filters['department_id'])

            # Filter by location
            if 'location_id' in filters:
                queryset = queryset.filter(location_id=filters['location_id'])

        # Create workbook
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = 'Assets'

        # Define header style
        header_font = Font(bold=True, size=11)
        header_fill = PatternFill(start_color='CCE5FF', end_color='CCE5FF', fill_type='solid')
        header_alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )

        # Write headers
        for col_idx, column in enumerate(columns, 1):
            cell = ws.cell(row=1, column=col_idx)
            cell.value = column_headers.get(column, column)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
            cell.border = border

        # Write data rows
        row_idx = 2
        for asset in queryset:
            for col_idx, column in enumerate(columns, 1):
                cell = ws.cell(row=row_idx, column=col_idx)
                cell.border = border

                # Get value based on column
                value = ''
                if column == 'asset_code':
                    value = asset.asset_code or ''
                elif column == 'asset_name':
                    value = asset.asset_name or ''
                elif column == 'specification':
                    value = asset.specification or ''
                elif column == 'brand':
                    value = asset.brand or ''
                elif column == 'model':
                    value = asset.model or ''
                elif column == 'serial_number':
                    value = asset.serial_number or ''
                elif column == 'asset_category':
                    value = asset.asset_category.name if asset.asset_category else ''
                elif column == 'supplier':
                    value = asset.supplier.supplier_name if asset.supplier else ''
                elif column == 'department':
                    value = asset.department.department_name if asset.department else ''
                elif column == 'location':
                    value = asset.location.full_name if hasattr(asset.location, 'full_name') else (asset.location.location_name if asset.location else '')
                elif column == 'custodian':
                    value = asset.custodian.username if asset.custodian else ''
                elif column == 'asset_status':
                    value = asset.get_asset_status_display() or ''
                elif column == 'purchase_date':
                    value = asset.purchase_date.strftime('%Y-%m-%d') if asset.purchase_date else ''
                elif column == 'original_value':
                    value = float(asset.original_value) if asset.original_value else 0
                elif column == 'net_value':
                    value = float(asset.net_value) if asset.net_value else 0
                elif column == 'useful_life':
                    value = asset.useful_life or 0
                elif column == 'created_at':
                    value = asset.created_at.strftime('%Y-%m-%d %H:%M:%S') if asset.created_at else ''

                cell.value = value
                cell.alignment = Alignment(vertical='center')

            row_idx += 1

        # Auto-adjust column widths
        for col in ws.columns:
            max_length = 0
            column = col[0].column_letter
            for cell in col:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column].width = adjusted_width

        # Save to bytes buffer
        buffer = BytesIO()
        wb.save(buffer)
        buffer.seek(0)

        # Generate filename with timestamp
        timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
        filename = f'assets_export_{timestamp}.xlsx'

        # Return as file download
        response = HttpResponse(
            buffer.read(),
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = f'attachment; filename="{filename}"'

        return response


class SupplierViewSet(BaseModelViewSetWithBatch):
    """
    ViewSet for Supplier management.

    Provides standard CRUD operations for suppliers.
    """

    queryset = Supplier.objects.all()
    serializer_class = SupplierSerializer
    filterset_class = SupplierFilter
    service = SupplierService()

    def get_serializer_class(self):
        """Return appropriate serializer based on action."""
        if self.action == 'list':
            return SupplierListSerializer
        return super().get_serializer_class()

    def get_serializer_context(self):
        """Add organization_id to serializer context."""
        context = super().get_serializer_context()
        # Try to get organization_id from request attribute (set by middleware)
        if hasattr(self.request, 'organization_id') and self.request.organization_id:
            context['organization_id'] = str(self.request.organization_id)
        # Fallback to user's organization_id if available
        elif hasattr(self.request, 'user') and hasattr(self.request.user, 'organization_id'):
            context['organization_id'] = str(self.request.user.organization_id)
        return context

    def perform_create(self, serializer):
        """Set organization and created_by on create."""
        organization_id = getattr(self.request, 'organization_id', None)
        serializer.save(
            organization_id=organization_id,
            created_by=self.request.user
        )

    def create(self, request, *args, **kwargs):
        """Override to wrap response in standard format."""
        serializer = self.get_serializer(data=request.data)
        serializer.is_valid(raise_exception=True)
        self.perform_create(serializer)

        return BaseResponse.created(
            data=serializer.data,
            message='Supplier created successfully'
        )

    def retrieve(self, request, *args, **kwargs):
        """Override to wrap response in standard format."""
        instance = self.get_object()
        serializer = self.get_serializer(instance)
        return BaseResponse.success(
                data=serializer.data,
                message='Query successful'
            )

    def update(self, request, *args, **kwargs):
        """Override to wrap response in standard format."""
        partial = kwargs.pop('partial', False)
        instance = self.get_object()
        serializer = self.get_serializer(
            instance, data=request.data, partial=partial
        )
        serializer.is_valid(raise_exception=True)
        self.perform_update(serializer)

        return BaseResponse.success(
                data=serializer.data,
                message='Supplier updated successfully'
            )

    def destroy(self, request, *args, **kwargs):
        """Override to wrap response in standard format."""
        instance = self.get_object()
        self.perform_destroy(instance)
        return BaseResponse.success(message='Supplier deleted successfully')


class LocationViewSet(BaseModelViewSetWithBatch):
    """
    ViewSet for Location management.

    Provides:
    - Standard CRUD operations
    - Tree endpoint for hierarchical data
    """

    queryset = Location.objects.all()
    serializer_class = LocationSerializer
    filterset_class = LocationFilter
    service = LocationService()

    def get_serializer_class(self):
        """Return appropriate serializer based on action."""
        if self.action == 'list':
            return LocationListSerializer
        return super().get_serializer_class()

    def get_serializer_context(self):
        """Add organization_id to serializer context."""
        context = super().get_serializer_context()
        # Try to get organization_id from request attribute (set by middleware)
        if hasattr(self.request, 'organization_id') and self.request.organization_id:
            context['organization_id'] = str(self.request.organization_id)
        # Fallback to user's organization_id if available
        elif hasattr(self.request, 'user') and hasattr(self.request.user, 'organization_id'):
            context['organization_id'] = str(self.request.user.organization_id)
        return context

    def perform_create(self, serializer):
        """Set organization and created_by on create."""
        organization_id = getattr(self.request, 'organization_id', None)
        serializer.save(
            organization_id=organization_id,
            created_by=self.request.user
        )

    def create(self, request, *args, **kwargs):
        """Override to wrap response in standard format."""
        serializer = self.get_serializer(data=request.data)
        serializer.is_valid(raise_exception=True)
        self.perform_create(serializer)

        return BaseResponse.created(
            data=serializer.data,
            message='Location created successfully'
        )

    def retrieve(self, request, *args, **kwargs):
        """Override to wrap response in standard format."""
        instance = self.get_object()
        serializer = self.get_serializer(instance)
        return BaseResponse.success(
                data=serializer.data,
                message='Query successful'
            )

    def update(self, request, *args, **kwargs):
        """Override to wrap response in standard format."""
        partial = kwargs.pop('partial', False)
        instance = self.get_object()
        serializer = self.get_serializer(
            instance, data=request.data, partial=partial
        )
        serializer.is_valid(raise_exception=True)
        self.perform_update(serializer)

        return BaseResponse.success(
                data=serializer.data,
                message='Location updated successfully'
            )

    def destroy(self, request, *args, **kwargs):
        """Override to wrap response in standard format."""
        instance = self.get_object()
        self.perform_destroy(instance)
        return BaseResponse.success(message='Location deleted successfully')

    @action(detail=False, methods=['get'], url_path='tree')
    def tree(self, request):
        """
        Get complete location tree.

        GET /api/assets/locations/tree/
        """
        organization_id = getattr(request, 'organization_id', None)

        if not organization_id:
            return BaseResponse.error(
                    code='UNAUTHORIZED',
                    message='Organization context required'
                )

        tree_data = self.service.get_tree(organization_id)

        return BaseResponse.success(
                data=tree_data,
                message='Location tree retrieved successfully'
            )


class AssetStatusLogViewSet(BaseModelViewSetWithBatch):
    """
    ViewSet for Asset Status Log viewing.

    Status logs are read-only (created by status change actions).
    """

    queryset = AssetStatusLog.objects.select_related('asset', 'created_by').all()
    serializer_class = AssetStatusLogSerializer
    filterset_class = AssetStatusLogFilter

    def get_serializer_class(self):
        """Return appropriate serializer based on action."""
        if self.action == 'list':
            return AssetStatusLogListSerializer
        return super().get_serializer_class()

    def get_queryset(self):
        """Filter by organization through asset relation."""
        queryset = super().get_queryset()
        # Additional filtering can be added here if needed
        return queryset

    def retrieve(self, request, *args, **kwargs):
        """Override to wrap response in standard format."""
        instance = self.get_object()
        serializer = self.get_serializer(instance)
        return BaseResponse.success(
                data=serializer.data,
                message='Query successful'
            )

    def list(self, request, *args, **kwargs):
        """Override to wrap response in standard format."""
        queryset = self.filter_queryset(self.get_queryset())

        page = self.paginate_queryset(queryset)
        if page is not None:
            serializer = self.get_serializer(page, many=True)
            return self.get_paginated_response(serializer.data)

        serializer = self.get_serializer(queryset, many=True)
        return BaseResponse.success(
                data=serializer.data,
                message='Query successful'
            )
